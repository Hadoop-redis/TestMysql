#!/usr/bin/python3
# -*- coding:utf-8 -*-
"""
@author: tianwen
@file: test.py
@time: 2020/8/28 9:45
@desc: 操作excel的测试
"""
import locale
from openpyxl import Workbook
import time
import datetime

"""创建一个excel文件，并写入不同的内容"""
# # 创建文件对象
# wb = Workbook()
#
# # 获取第一个sheet
# ws = wb.active
# # 将数据写入到指定的单元格
# ws['A1'] = 42
# ws['B1'] = "自动化" + "automation test"
# ws['C1'] = "自动化automation test"
#
# # 写入多个单元格
# ws.append([1, 2, 3])
#
# # 保存为a.xlsx
# wb.save("a.xlsx")

"""写入时间"""
# # 创建对象
# wb = Workbook()
# # 获取第一个sheet
# ws = wb.active
# # 写入一个当前时间
# ws['A2'] = datetime.datetime.now()
#
# # 写入一个自定义的时间格式
# locale.setlocale(locale.LC_CTYPE, 'chinese')
# ws['A3'] = time.strftime("%Y年%m月%d日 %H时%M分%S秒", time.localtime())
#
# # svae
# wb.save("sample.xlsx")

"""创建sheet"""
wb = Workbook()

ws = wb.create_sheet('kongsh')
ws1 = wb.create_sheet("Mysheet")

ws1.title = "New Title"
# 设定sheet的插入位置
ws2 = wb.create_sheet("Mysheet", 0)
# 设置sheet名的背景色
ws1.sheet_properties.tabColor = "1072BA"

# 获取某个sheet对象
print(wb["测试用例"])
print(wb["New Title"])

# 获取所有的sheet名字
print(wb.sheetnames)
for sheet_name in wb.sheetnames:
    print(sheet_name)
    print(wb[sheet_name])

# 遍历获取sheet对象，按照sheet顺序获取
for sheet in wb:
    print(sheet)
    print(sheet.title)

# 复制一个sheet
wb["New Title"]["A1"] = "gloryroad"
source = wb["New Title"]
target = wb.copy_worksheet(source)

target.title = "New copy Title"

# save
wb.save("b.xlsx")
